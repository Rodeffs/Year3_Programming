import openpyxl
from openpyxl.chart import PieChart, Reference 


def main():
    wb = openpyxl.load_workbook("salaries.xlsx")
    sheet = wb.active

    sheet.cell(2, 11).value = "Бухгалтерия"
    sheet.cell(2, 12).value = sheet.cell(4,9).value

    sheet.cell(3, 11).value = "Отдел кадров"
    sheet.cell(3, 12).value = sheet.cell(9,9).value

    sheet.cell(4, 11).value = "Столовая"
    sheet.cell(4, 12).value = sheet.cell(11,9).value

    pie = PieChart() 
   
    labels = Reference(sheet, min_col=11, min_row=2, max_row=4)
    data = Reference(sheet, min_col=12, min_row=2, max_row=4)

    pie.add_data(data)
    pie.set_categories(labels)
    pie.title = "Зарплата по отделам"

    sheet.add_chart(pie, 'M1')

    wb.save("salaries.xlsx")


if __name__ == "__main__":
    main()

